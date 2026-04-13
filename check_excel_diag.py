import openpyxl
import os

BASE_DIR = r"a:\TOOLS\kodlama\km\PTS"
EYLEM_XLSX = os.path.join(BASE_DIR, "OÖP OKUL EYLEM PLANI.xlsx")
TAKTAK_XLSX = os.path.join(BASE_DIR, "OGP FAALİYET TAKVİMİ.xlsx")

def run_check():
    try:
        if os.path.exists(TAKTAK_XLSX):
            print(f"Excel Takvim Kontrolü: {TAKTAK_XLSX}")
            wb = openpyxl.load_workbook(TAKTAK_XLSX, data_only=True)
            ws = wb.active
            
            # Renk örneklemesi
            colored_cells = []
            for r in range(1, 15):
                for c in range(1, 40):
                    fill = ws.cell(r, c).fill
                    if fill and fill.start_color.index != '00000000':
                        val = ws.cell(r, c).value
                        colored_cells.append(f"R{r}C{c}: [{val}] Color:{fill.start_color.index}")
            
            print(f"Bulunan renkli hücreler (ilk 15 satır):")
            for cc in colored_cells[:20]:
                print(cc)
            
            print(f"\nHeader Örneği (R1-R3):")
            for r in range(1, 4):
                line = [str(ws.cell(r, c).value) for c in range(1, 6)]
                print(f"R{r}: {line}")
        else:
            print("TAKTAK_XLSX bulunamadı.")
    except Exception as e:
        print(f"HATA: {e}")

if __name__ == "__main__":
    run_check()
