import openpyxl, os
base = r"a:\TOOLS\kodlama\km\PT SİHİRBAZI"
for fname in ["ogp proiz rapor.xlsx","OGP FAALİYET TAKVİMİ.xlsx","OGP OKUL EYLEM PLANI.xlsx","OÖP FAALİYET TAKVİMİ.xlsx","OÖP OKUL EYLEM PLANI.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(base, fname), data_only=True)
    print(f"\n=== {fname} ===")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  Sayfa: {sn} ({ws.max_row}x{ws.max_column})")
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if i>=15: break
            if any(c is not None for c in row): print(f"    {i+1}: {list(row)}")
input("Bitti. Enter'a basin.")
