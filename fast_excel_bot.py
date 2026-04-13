import openpyxl
import os
import re
import calendar
import unicodedata

# Dosya yollari
BASE_DIR = r"a:\TOOLS\kodlama\km\PTS"
EYLEM_XLSX = os.path.join(BASE_DIR, "OÖP OKUL EYLEM PLANI.xlsx")
TAKTAK_XLSX = os.path.join(BASE_DIR, "OGP FAALİYET TAKVİMİ.xlsx")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "Eylem_Plani_Raporu.xlsx")

turkish_months = {
    "EYLUL": 9, "EKIM": 10, "KASIM": 11, "ARALIK": 12,
    "OCAK": 1, "SUBAT": 2, "MART": 3, "NISAN": 4,
    "MAYIS": 5, "HAZIRAN": 6, "TEMMUZ": 7, "AGUSTOS": 8
}

def clean_text(text):
    if text is None: return ""
    return str(text).strip()

def clean_task_name(name):
    name = clean_text(name)
    # Temizleme: "1. ", "Eylem 1:", "Görev 2:" vb.
    name = re.sub(r'^(Eylem|Görev|Sıra)?\s*\d+[\.\s:]*', '', name, flags=re.I).strip()
    return name

def universal_key(text):
    if not text: return ""
    text = text.lower().replace('ı', 'i').replace('İ', 'i')
    # Aksanları temizle (ü -> u, ş -> s vb.)
    text = ''.join(c for c in unicodedata.normalize('NFKD', text) if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z0-9]', '', text)

def run_extraction():
    print("Excel-tabanlı işlem başlatıldı...")
    
    tasks_data = []

    # 1. EYLEM PLANI OKUMA
    try:
        print(f"Eylem Planı okunuyor: {EYLEM_XLSX}")
        wb_e = openpyxl.load_workbook(EYLEM_XLSX, data_only=True)
        ws_e = wb_e.active
        
        # OÖP dosyasında başlığı bulalım.
        for r in range(1, 100): # İlk 100 satır
            for c in range(1, 10): # İlk 10 sütun
                val = clean_text(ws_e.cell(r, c).value)
                if "Eylemler" in val and "Görevler" in val:
                    # Başlık bulundu! r. satır, c. sütun
                    task_col = c
                    data_col = c + 3 # Sorumlu genelde 3 yan sütundadır (C5 - C2 = 3)
                    
                    for r_idx in range(r + 1, ws_e.max_row + 1):
                        raw_name = ws_e.cell(r_idx, task_col).value
                        if not raw_name: continue
                        
                        name = clean_task_name(raw_name)
                        col4 = clean_text(ws_e.cell(r_idx, data_col).value)
                        
                        tasks_data.append({
                            "name": clean_text(raw_name),
                            "match_key": universal_key(name),
                            "col4": col4,
                            "dates": []
                        })
                    break
            if tasks_data: break
        print(f"{len(tasks_data)} görev toplandı.")
    except Exception as e:
        print(f"Eylem Planı Okuma Hatası: {e}")
        return

    # 2. TAKVİM ANALİZİ
    try:
        print(f"Takvim analiz ediliyor: {TAKTAK_XLSX}")
        wb_t = openpyxl.load_workbook(TAKTAK_XLSX, data_only=True)
        ws_t = wb_t.active
        
        # Header Mapping (Yıl ve Ay)
        col_to_date = {} # col_idx -> (year, month)
        
        # Yılları bul (Row 3'te)
        years = {}
        curr_y = None
        for c in range(1, ws_t.max_column + 1):
            v = ws_t.cell(3, c).value
            if v and str(v).isdigit(): curr_y = int(v)
            if curr_y: years[c] = curr_y
            
        # Ayları bul (Row 4'te)
        for c in range(1, ws_t.max_column + 1):
            v = clean_text(ws_t.cell(4, c).value)
            m_num = None
            if v:
                m_norm = universal_key(v).upper()
                for mname, mnum in turkish_months.items():
                    if mname in m_norm: m_num = mnum; break
            
            if m_num and years.get(c):
                col_to_date[c] = (years[c], m_num)

        # Görev Eşleştirme ve Renk Tespiti
        # Takvimde görevler C1 veya C2'dedir.
        match_count = 0
        for r in range(5, ws_t.max_row + 1):
            c1_val = clean_text(ws_t.cell(r, 1).value)
            c2_val = clean_text(ws_t.cell(r, 2).value)
            
            row_key = universal_key(clean_task_name(c1_val or c2_val))
            if not row_key: continue
            
            target_task = None
            for t in tasks_data:
                if t["match_key"] and (t["match_key"] in row_key or row_key in t["match_key"]):
                    target_task = t
                    break
            
            if target_task:
                match_count += 1
                # Renkli hücreleri topla
                green_cols = []
                for c in sorted(col_to_date.keys()):
                    fill = ws_t.cell(r, c).fill
                    # Renk var mı (Beyaz/Boş değilse)
                    if fill and fill.start_color.index != '00000000':
                        green_cols.append(c)
                
                if green_cols:
                    # Gruplandır (Ardışık sütunları tek tarih aralığı yap)
                    groups, curr_g = [], [green_cols[0]]
                    for i in range(1, len(green_cols)):
                        if green_cols[i] == green_cols[i-1] + 1: curr_g.append(green_cols[i])
                        else: groups.append(curr_g); curr_g = [green_cols[i]]
                    groups.append(curr_g)
                    
                    for g in groups:
                        sy, sm = col_to_date[g[0]]
                        ey, em = col_to_date[g[-1]]
                        sd = f"01.{sm:02d}.{sy}"
                        ed = f"{calendar.monthrange(ey, em)[1]:02d}.{em:02d}.{ey}"
                        target_task["dates"].append((sd, ed))

        print(f"{match_count} görev eşleşti.")
    except Exception as e:
        print(f"Takvim Analiz Hatası: {e}")

    # 3. EXCEL RAPORU OLUŞTURMA
    try:
        print(f"Rapor oluşturuluyor: {OUTPUT_EXCEL}")
        wb_o = openpyxl.Workbook()
        ws_o = wb_o.active
        ws_o.append(["Sıra", "Eylem / Görev", "Tekrar", "", "", "Başlangıç 1", "Bitiş 1", "Başlangıç 2", "Bitiş 2", "Başlangıç 3", "Bitiş 3", "Başlangıç 4", "Bitiş 4", "İlgili Veri"])
        
        for i, t in enumerate(tasks_data, 1):
            row = [i, t["name"], len(t["dates"]), "", ""]
            for d in t["dates"][:4]: row.extend([d[0], d[1]])
            while len(row) < 13: row.append("")
            row.append(t["col4"])
            ws_o.append(row)
        
        wb_o.save(OUTPUT_EXCEL)
        print("[OK] İşlem saniyeler içinde başarıyla tamamlandı!")
    except Exception as e:
        print(f"Rapor Yazma Hatası: {e}")

if __name__ == "__main__":
    run_extraction()
