import openpyxl
import os

BASE_DIR = r"a:\TOOLS\kodlama\km\PTS"
TAKTAK_XLSX = os.path.join(BASE_DIR, "OGP FAALİYET TAKVİMİ.xlsx")

def run_check():
    wb = openpyxl.load_workbook(TAKTAK_XLSX, data_only=True)
    ws = wb.active
    
    # Yıl başlıkları (R3)
    years = {}
    for c in range(1, 100):
        v = ws.cell(3, c).value
        if v and str(v).isdigit():
            years[c] = int(v)
    
    # Ay başlıkları (R4)
    months = {}
    for c in range(1, 100):
        v = ws.cell(4, c).value
        if v:
            months[c] = str(v).strip()

    print(f"Yıllar (R3): {years}")
    print(f"Aylar (R4): {months}")

if __name__ == "__main__":
    run_check()
