import openpyxl
import os

BASE_DIR = r"a:\TOOLS\kodlama\km\PTS"
TAKTAK_XLSX = os.path.join(BASE_DIR, "OGP FAALİYET TAKVİMİ.xlsx")

def run_check():
    wb = openpyxl.load_workbook(TAKTAK_XLSX, data_only=True)
    ws = wb.active
    
    print("Excel Takvim Analizi (Aktivite ve Renk Bulma):")
    for r in range(1, 30):
        c1 = str(ws.cell(r,1).value or "").strip()
        c2 = str(ws.cell(r,2).value or "").strip()
        
        # Eğer bir aktivite ismi bulursak (Uzun metin veya "1. ", "Eylem" vb.)
        if len(c1) > 10 or len(c2) > 10:
            row_colors = []
            for c in range(3, 30):
                fill = ws.cell(r, c).fill
                if fill and fill.start_color.index != '00000000':
                    row_colors.append(f"C{c}:{fill.start_color.index}")
            
            if row_colors:
                print(f"R{r} (C1: {c1[:20]}, C2: {c2[:20]}): {row_colors}")

if __name__ == "__main__":
    run_check()
