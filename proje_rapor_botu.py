import openpyxl
import re
import os
from openpyxl.styles import PatternFill
import calendar

# Klasör yolu
BASE = r"a:\TOOLS\kodlama\km\PT SİHİRBAZI"

def clean_activity_name(name):
    if not name: return ""
    name = re.sub(r'^\d+\.\s*', '', str(name))
    if "(Hedef" in name:
        name = name.split("(Hedef")[0]
    return name.strip()

def get_gray_blocks(row):
    """Gri hücre bloklarını ve hangi ay sütununda olduklarını döner."""
    blocks = []
    in_block = False
    
    # Takvim sütunları 3. sütundan (C) başlıyor varsayalım (Eylül, Ekim...)
    for col_idx, cell in enumerate(row[2:], 3): 
        fill = cell.fill
        is_gray = False
        
        # Herhangi bir dolgu rengini 'işaretli' kabul et
        if fill and fill.start_color and fill.start_color.index != '00000000':
            is_gray = True
            
        if is_gray and not in_block:
            blocks.append(col_idx) # Blok başlangıç sütunu
            in_block = True
        elif not is_gray and in_block:
            in_block = False
            
    return blocks

def calculate_dates(start_columns):
    """Sütun indekslerine göre Y1, Y2... tarihlerini oluşturur."""
    # Varsayım: Sütun 3 = Eylül (9.ay), 4 = Ekim (10.ay)...
    # 2025-2026 Eğitim Yılı
    results = []
    
    for i, col in enumerate(start_columns, 1):
        # Sütun 3 ise ay 9, sütun 4 ise ay 10...
        month = (col - 3 + 9)
        year = 2025
        if month > 12:
            month -= 12
            year = 2026
            
        last_day = calendar.monthrange(year, month)[1]
        results.append(f"Y{i} başlangıç 01.{month:02d}.{year} Y{i} son {last_day:02d}.{month:02d}.{year}")
        
    return ", ".join(results)

def run_process():
    print("OÖP Raporu oluşturma işlemi başlatıldı...")
    
    try:
        # data_only=True değerler için, False ise renkleri okumak için
        wb_eylem = openpyxl.load_workbook(os.path.join(BASE, "OÖP OKUL EYLEM PLANI.xlsx"), data_only=True)
        wb_takvim = openpyxl.load_workbook(os.path.join(BASE, "OÖP FAALİYET TAKVİMİ.xlsx"))
        wb_template = openpyxl.load_workbook(os.path.join(BASE, "ogp proiz rapor.xlsx"))
    except Exception as e:
        print(f"HATA: Dosyalar okunurken bir sorun oluştu: {e}")
        return

    ws_eylem = wb_eylem.active
    ws_takvim = wb_takvim.active
    ws_rapor = wb_template.active
    
    # Takvim haritası oluştur (İsim bazlı)
    takvim_map = {}
    for row in ws_takvim.iter_rows(min_row=2):
        name = clean_activity_name(row[1].value)
        if name:
            takvim_map[name] = row

    # Raporu temizle ve doldur (Sütunlar: B=İsim, D=Tarih, E=Sorumlu)
    # Şabonun 2. satırdan başladığını varsayıyoruz
    report_row = 2
    
    for row in ws_eylem.iter_rows(min_row=2):
        raw_name = row[1].value
        if not raw_name: continue
        
        clean_name = clean_activity_name(raw_name)
        # Eylem Planı E sütunu (Sorumlu) - 5. sütun
        responsible = str(row[4].value if row[4].value else "")
        
        blocks = []
        if clean_name in takvim_map:
            blocks = get_gray_blocks(takvim_map[clean_name])
        
        # Tarih ve Tekrar hesapla
        date_str = calculate_dates(blocks) if blocks else "Tarih Belirtilmemiş"
        rep_count = len(blocks) - 1 if len(blocks) > 0 else 0
        
        # Rapora yaz
        ws_rapor.cell(report_row, 2).value = clean_name
        ws_rapor.cell(report_row, 4).value = date_str  # D Sütunu: Tarihler
        ws_rapor.cell(report_row, 5).value = responsible # E Sütunu: Sorumlu
        
        print(f"Tamamlandı: {clean_name}")
        report_row += 1

    # Kaydet
    try:
        out_path = os.path.join(BASE, "oöp proiz rapor.xlsx")
        wb_template.save(out_path)
        print(f"\n✅ İŞLEM BAŞARILI!")
        print(f"Oluşturulan dosya: {out_path}")
    except Exception as e:
        print(f"KAYDETME HATASI: {e} (Dosya açık olabilir!)")

if __name__ == "__main__":
    run_process()
